#!/usr/bin/env python3
"""生成标准发送任务表格模版（带样式）"""
import sys, os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def generate_template(save_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 发送任务 sheet ─────────────────────────────────
    ws = wb.create_sheet('发送任务')
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 28
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 10

    # Row 1: 合并标题行
    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = '🦐 微信批量发送任务表  |  填写说明：带 * 号列为必填；发送时间留空 = 立即发送'
    c.font = Font(bold=True, color='FFFFFF', size=12)
    c.fill = PatternFill('solid', fgColor='3D50CC')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Row 2: 表头
    headers = ['#', '* 应用', '* 联系人/群聊', '* 消息类型', '* 文字内容',
                '图片路径', '发送时间', '重复', '备注', '状态']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(2, ci, h)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', fgColor='1A1D27')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 24

    # Row 3: 示例数据
    ws.row_dimensions[3].height = 20
    example = ['1', '微信', '万旗', '文字', '这是一条自动发送的消息',
               '', '', '', '例：群发通知测试', '发送成功 03-24 17:05']
    for ci, v in enumerate(example, 1):
        cell = ws.cell(3, ci, v)
        cell.alignment = Alignment(vertical='center')
        if ci == 10:
            cell.fill = PatternFill('solid', fgColor='FFF8D7')
            cell.font = Font(color='8B6914')

    # ── 填写说明 sheet ─────────────────────────────────
    ws2 = wb.create_sheet('填写说明')
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 44
    ws2.column_dimensions['C'].width = 40

    for ci, h in enumerate(['字段', '说明', '示例'], 1):
        cell = ws2.cell(1, ci, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='E8F9F0')

    guide = [
        ('应用', '微信 / 钉钉 / 飞书（三选一）', '微信'),
        ('联系人/群聊', '精确的联系人昵称或群聊名称（用于搜索）', '产品讨论群'),
        ('消息类型', '文字 / 图片 / 文字+图片', '文字+图片'),
        ('文字内容', '支持变量：{name} {date} {time}', '你好 {name}，今日报告请查收'),
        ('图片路径', '本机绝对路径，支持 jpg/png/gif', '/Users/you/pic.jpg'),
        ('发送时间', '留空=立即发送；格式 YYYY-MM-DD HH:MM', '2026-03-24 15:00'),
        ('重复', '留空=单次；daily=每天；weekly=每周；workday=工作日', 'daily'),
        ('备注', '仅供人类阅读，不影响发送', 'Q1 通知'),
        ('状态', '由程序自动填入，无需手动填写', '已完成'),
    ]
    for ri, (field, desc, ex) in enumerate(guide, 2):
        ws2.cell(ri, 1, field).font = Font(bold=True)
        ws2.cell(ri, 2, desc)
        ws2.cell(ri, 3, ex).font = Font(color='07C160')

    wb.save(save_path)
    print(f'OK:{save_path}')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('ERROR: missing save_path argument')
        sys.exit(1)
    generate_template(sys.argv[1])
